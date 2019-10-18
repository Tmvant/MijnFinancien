# -*- coding: utf-8 -*-
"""
Created on Tue Oct  3 14:57:48 2017

@author: tmvant
"""

#TODO: feature om alles te reprocessen
#https://forum.knime.com/t/datavalidation-drop-down-list-in-excel-using-python-scrpt-node/12648

import os, sys
import configparser
import argparse

import numpy as np
import pandas as pd
import openpyxl as pyxl
from openpyxl.worksheet.datavalidation import DataValidation

from rules import rules

PATH_INBOX = 'Inbox'
PATH_OVERVIEWS = 'Overzichten'
PATH_SETTINGS = 'Instellingen'

def insert_data_validation(file, categories):
    wb = pyxl.load_workbook(file)
    wb.create_sheet('Categorieen')
    ws = wb['Categorieen']
    ws.append(categories)
    ws = wb['Sheet1']
    ws.title = 'Transacties'
    category_dv = DataValidation(type='list',formula1='=Categorieen!$1:$1')
    ws.add_data_validation(category_dv)
    _ = [category_dv.add('B' + str(i)) for i in range(2,len(ws['A'])+1)]
    wb.save(file)    

def get_files(path):
    return [os.path.join(path, f) for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]

def reverse_date(date):
    parts = date.split('-')
    return f"{parts[2]}-{parts[1]}-{parts[0]}"


def main(args):
    overview_file = os.path.join(PATH_OVERVIEWS,f"{args.year}.xlsx")

    banks = configparser.ConfigParser()
    banks.read(os.path.join(PATH_SETTINGS,'bankdocumenten.ini'))

    accounts = configparser.ConfigParser()
    accounts.read(os.path.join(PATH_SETTINGS,'rekeningen.ini'))

    overview_wb = pyxl.load_workbook(overview_file)
    
    # Load categories from sheet
    categories = []
    for category in overview_wb['Categorieen']['A']:
        categories.append(category.value)
    categories = [x for x in categories if x is not None]

    writer = pd.ExcelWriter(overview_file, engine='openpyxl') 
    writer.book = overview_wb
    writer.sheets = dict((ws.title, ws) for ws in overview_wb.worksheets)

    # Load files from inbox into dataframe
    inbox_files = get_files(PATH_INBOX)
    entries = pd.DataFrame()
    for file in inbox_files:
        entries = entries.append(pd.read_csv(file, index_col=False, names=banks[args.bank]['FormaatCSV'].split(',')))

    # Select relevant columns
    entries = entries[['Date','Account','AccountOther','NameOther','Amount','JournalDate','Type','Id','Description']]
    
    # Add categories column
    entries.insert(0,column='Category',value=[np.nan] * len(entries['Date']))

    # Add "New" column
    entries.insert(0,column='New',value=['X'] * len(entries['Date']))

    # Reformat date (from 01-10-2000 to 2000-10-01)
    entries['Date'] = entries['Date'].apply(reverse_date)
    entries['JournalDate'] = entries['JournalDate'].apply(reverse_date)

    # Filter by current year
    entries = entries[entries['Date'].str.match(args.year)]

    # Load existing transactions
    transactions = pd.DataFrame()
    for account in accounts.keys():
        sheet_name = account + " (Tr)"
        if sheet_name in overview_wb.sheetnames:
            data = pd.read_excel(overview_file, sheet_name=sheet_name)
            if not data.empty:
                transactions = transactions.append(pd.read_excel(writer, sheet_name=sheet_name, index_col=0))
    if transactions.empty:
        transactions = entries
    else:
        # Empty 'Nieuw' column
        transactions['New'] = ''

        # Append new entries
        transactions = transactions.append(entries)

    # Get unique entries
    unique_key = banks[args.bank]['UniekeSleutel'].split(',')
    transactions = transactions.drop_duplicates(subset=unique_key, keep='first')

    # Select categoryless and new entries
    to_process = transactions[transactions['Category'].isnull()]
    to_process = to_process[to_process['New'] == 'X']
    if not to_process.empty:
        # Auto-categorize
        for category in categories:
            if category not in rules.keys():
                continue
            category_rules = rules[category]
            for rule in category_rules:
                to_process['Category'].iloc[np.where(rule(to_process))] = category

    # Merge newly categorized
    transactions = transactions.append(to_process)
    transactions = transactions.drop_duplicates(subset=unique_key, keep='last')

    # Sort by date
    transactions = transactions.sort_values(by=['Date'])

    # group by account
    account_frames = [f for f in transactions.groupby(['Account'])]

    # Save to xlsx
    for frame in account_frames:
        sheet_name = ''
        account_number = frame[0]
        for account in accounts.keys():
            if account is 'DEFAULT':
                continue
            if account_number == accounts[account]['IBAN']:
                sheet_name = account + " (Tr)"
                break
        if sheet_name is not '':
            frame[1].to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()

    # Reload xlsx and add datavalidation options
    #intermediate_files = get_files(PATH_INTERMEDIATE)
    #for file in intermediate_files:
    #    insert_data_validation(file, categories)

    overview_wb.close()

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-b', '--bank', type=str, default='')
    parser.add_argument('-y', '--year', type=str, default='')
    args = parser.parse_args()
    main(args)